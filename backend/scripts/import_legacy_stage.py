#!/usr/bin/env python
"""Import legacy XLSX files into staging tables (legacy_*)."""
from __future__ import annotations

import argparse
import logging
import os
import re
from pathlib import Path
from typing import Iterable, List

import django
from django.db import connection, transaction

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required to run this script. Install it with: pip install openpyxl"
    ) from exc


LOGGER = logging.getLogger(__name__)


def normalize_identifier(name: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_]+", "_", name.strip())
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned.lower() or "column"


def build_unique_columns(headers: Iterable[str]) -> List[str]:
    seen = {}
    unique = []
    for idx, raw in enumerate(headers, start=1):
        base = normalize_identifier(str(raw) if raw is not None else f"col_{idx}")
        count = seen.get(base, 0)
        if count:
            col = f"{base}_{count + 1}"
        else:
            col = base
        seen[base] = count + 1
        unique.append(col)
    return unique


def table_name_for_file(path: Path) -> str:
    return f"legacy_{normalize_identifier(path.stem)}"


def list_xlsx_files(data_dir: Path) -> List[Path]:
    files = sorted(path for path in data_dir.glob("*.xlsx"))
    return [path for path in files if not path.name.lower().startswith("dbo_view_")]


def get_existing_columns(table: str) -> List[str]:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = %s
            ORDER BY ordinal_position
            """,
            [table],
        )
        return [row[0] for row in cursor.fetchall()]


def ensure_table(table: str, columns: List[str]) -> None:
    quote = connection.ops.quote_name
    existing = get_existing_columns(table)
    if not existing:
        columns_sql = ", ".join(f"{quote(col)} TEXT" for col in columns)
        with connection.cursor() as cursor:
            cursor.execute(
                f"CREATE TABLE IF NOT EXISTS {quote(table)} ({columns_sql});"
            )
        LOGGER.info("Created staging table %s with %s columns", table, len(columns))
        return

    missing = [col for col in columns if col not in existing]
    if missing:
        with connection.cursor() as cursor:
            for col in missing:
                cursor.execute(
                    f"ALTER TABLE {quote(table)} ADD COLUMN {quote(col)} TEXT;"
                )
        LOGGER.info("Added %s new columns to %s", len(missing), table)


def clear_table(table: str) -> int:
    quote = connection.ops.quote_name
    with connection.cursor() as cursor:
        cursor.execute(f"DELETE FROM {quote(table)};")
        return cursor.rowcount


def insert_rows(table: str, columns: List[str], rows: Iterable[List[object]], batch: int = 500) -> int:
    quote = connection.ops.quote_name
    col_sql = ", ".join(quote(col) for col in columns)
    placeholders = ", ".join(["%s"] * len(columns))
    sql = f"INSERT INTO {quote(table)} ({col_sql}) VALUES ({placeholders})"

    inserted = 0
    batch_rows: List[List[object]] = []

    with connection.cursor() as cursor:
        for row in rows:
            batch_rows.append(row)
            if len(batch_rows) >= batch:
                cursor.executemany(sql, batch_rows)
                inserted += len(batch_rows)
                batch_rows = []
        if batch_rows:
            cursor.executemany(sql, batch_rows)
            inserted += len(batch_rows)

    return inserted


def load_workbook_rows(path: Path) -> tuple[List[str], Iterable[List[object]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return [], []
    columns = build_unique_columns(headers)

    def row_iter() -> Iterable[List[object]]:
        for row in rows:
            normalized = ["" if value is None else str(value) for value in row]
            if len(normalized) < len(columns):
                normalized.extend([""] * (len(columns) - len(normalized)))
            yield normalized[: len(columns)]

    return columns, row_iter()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=Path(__file__).resolve().parents[2] / "data",
        help="Ruta al directorio data/ con XLSX legacy",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="Aplica cambios en la base de datos",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Simula la carga (default)",
    )
    return parser.parse_args()


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
    args = parse_args()
    dry_run = not args.commit or args.dry_run

    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
    django.setup()

    data_dir = args.data_dir
    if not data_dir.exists():
        raise SystemExit(f"No existe data dir: {data_dir}")

    files = list_xlsx_files(data_dir)
    if not files:
        LOGGER.warning("No se encontraron archivos XLSX en %s", data_dir)
        return

    total_inserted = 0
    with transaction.atomic():
        for path in files:
            LOGGER.info("Procesando %s", path.name)
            columns, row_iter = load_workbook_rows(path)
            if not columns:
                LOGGER.warning("Archivo vacío o sin encabezados: %s", path.name)
                continue

            table = table_name_for_file(path)
            ensure_table(table, columns)
            deleted = clear_table(table)
            if deleted:
                LOGGER.info("Limpiadas %s filas existentes en %s", deleted, table)

            inserted = insert_rows(table, columns, row_iter)
            total_inserted += inserted
            LOGGER.info("Insertadas %s filas en %s", inserted, table)

        LOGGER.info("Total filas insertadas en staging: %s", total_inserted)
        if dry_run:
            LOGGER.warning("Dry-run activo: realizando rollback")
            transaction.set_rollback(True)


if __name__ == "__main__":
    main()
