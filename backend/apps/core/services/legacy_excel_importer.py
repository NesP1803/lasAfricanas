from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.conf import settings
from django.contrib.auth import get_user_model
from django.db import connection
from django.db.models import Count
from django.utils import timezone
from openpyxl import load_workbook

from apps.core.models import Auditoria, ConfiguracionEmpresa, ConfiguracionFacturacion, Impuesto
from apps.inventario.models import Categoria, MovimientoInventario, Producto, Proveedor
from apps.taller.models import Mecanico, Moto
from apps.usuarios.models import PerfilVendedor
from apps.ventas.models import Cliente, DetalleVenta, RemisionAnulada, Venta, VentaAnulada


DOC_KEYS = ["documento", "cedula", "nit", "identificacion", "numero_documento", "idcliente", "clienteid"]
NAME_KEYS = ["nombre", "cliente", "razon_social", "razonsocial"]


def slug(text: Any) -> str:
    value = unicodedata.normalize("NFKD", str(text or "")).encode("ascii", "ignore").decode("ascii")
    value = re.sub(r"[\s\-.]+", "_", value.strip().lower())
    value = re.sub(r"[^a-z0-9_]+", "", value)
    return re.sub(r"_+", "_", value).strip("_")


def normalize_header(text: Any) -> str:
    base = slug(text)
    aliases = {
        "idcliente": "numero_documento",
        "clienteid": "numero_documento",
        "doc": "documento",
        "no_factura": "nofactura",
        "no_remision": "noremision",
        "num_factura": "nofactura",
        "num_remision": "noremision",
        "telefono_1": "telefono",
        "correo_electronico": "email",
    }
    return aliases.get(base, base)


def to_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    try:
        if value is None:
            return default
        if isinstance(value, bool):
            return default
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))

        text = str(value).strip()
        if not text or text.lower() in {"nan", "none", "null", "n/a", "na", "s/n", "-"}:
            return default

        text = text.replace(" ", "").replace("\u00a0", "")
        if text.endswith("%"):
            text = text[:-1]

        # limpieza de símbolos y ruido legacy sin eliminar separadores decimales
        text = re.sub(r"[^0-9,.\-]", "", text)
        if text in {"", "-", ".", ","}:
            return default

        if "." in text and "," in text:
            # Escenario 1.234,56 o 1,234.56
            text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")

        return Decimal(text)
    except (InvalidOperation, ValueError, TypeError):
        return default


def to_dt(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        dt = value
        if settings.USE_TZ and timezone.is_naive(dt):
            return timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    if isinstance(value, (int, float)):
        try:
            # Excel serial date base
            dt = datetime(1899, 12, 30) + timedelta(days=float(value))
            if settings.USE_TZ and timezone.is_naive(dt):
                return timezone.make_aware(dt, timezone.get_current_timezone())
            return dt
        except (TypeError, ValueError):
            return None
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            dt = datetime.strptime(text, fmt)
            if settings.USE_TZ and timezone.is_naive(dt):
                return timezone.make_aware(dt, timezone.get_current_timezone())
            return dt
        except ValueError:
            continue
    return None


def clean_value(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in {"", "nan", "none", "null", "n/a", "na", "s/n", "-"}:
            return None
    return value


@dataclass
class Dataset:
    path: Path
    sheet: str
    headers: list[str]
    rows: list[dict[str, Any]]
    raw_headers: list[str]


@dataclass
class FileReport:
    filename: str
    sheet: str
    classification: str
    rows_read: int = 0
    imported: int = 0
    updated: int = 0
    rejected: int = 0
    ambiguous: int = 0
    unmapped_columns: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)


class LegacyExcelImporter:
    def __init__(self, base_path: Path, commit: bool, cleanup_temp_on_success: bool):
        self.base_path = base_path
        self.commit = commit
        self.cleanup_temp_on_success = cleanup_temp_on_success
        self.reports: list[FileReport] = []
        self.unmapped_payloads: list[dict[str, Any]] = []
        self.file_classification: dict[str, str] = {}
        self.cache: dict[str, dict[str, Any]] = defaultdict(dict)

    def run(self) -> dict[str, Any]:
        datasets = self._load_datasets()
        for dataset in datasets:
            klass = self._classify(dataset)
            fr = FileReport(
                filename=dataset.path.name,
                sheet=dataset.sheet,
                classification=klass,
                rows_read=len(dataset.rows),
            )
            self.file_classification[dataset.path.name] = klass
            self._dispatch(dataset, fr)
            self.reports.append(fr)

        validations = self._run_validations()
        cleanup = self._cleanup_staging_tables(validations)
        report_payload = {
            "timestamp": timezone.now().isoformat(),
            "mode": "commit" if self.commit else "dry-run",
            "path": str(self.base_path),
            "files": [r.__dict__ for r in self.reports],
            "validations": validations,
            "unmapped_payloads": self.unmapped_payloads,
            "cleanup": cleanup,
        }
        return self._persist_reports(report_payload)

    def _load_datasets(self) -> list[Dataset]:
        datasets: list[Dataset] = []
        for path in sorted(self.base_path.glob("*.xlsx")):
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            row_iter = ws.iter_rows(values_only=True)
            try:
                raw_headers = [str(h or "") for h in next(row_iter)]
            except StopIteration:
                continue

            normalized: list[str] = []
            counts: dict[str, int] = defaultdict(int)
            for raw in raw_headers:
                key = normalize_header(raw)
                counts[key] += 1
                normalized.append(key if counts[key] == 1 else f"{key}__dup{counts[key]}")

            rows: list[dict[str, Any]] = []
            for row_no, row_values in enumerate(row_iter, start=2):
                cleaned = [clean_value(v) for v in row_values]
                if all(v in (None, "") for v in cleaned):
                    continue
                row = {normalized[i]: cleaned[i] if i < len(cleaned) else None for i in range(len(normalized))}
                row["_row_number"] = row_no
                rows.append(row)

            datasets.append(Dataset(path=path, sheet=ws.title, headers=normalized, raw_headers=raw_headers, rows=rows))
        return datasets

    def _classify(self, dataset: Dataset) -> str:
        h = set(dataset.headers)
        sample = dataset.rows[:8]
        scores = {
            "categorias": 3 * len(h & {"categoria", "nombre_categoria"}) + len(h & {"descripcion", "detalle"}),
            "impuestos": 3 * len(h & {"impuesto", "iva", "porcentaje"}),
            "productos": 4 * len(h & {"codigo", "articulo", "precio_venta", "costo", "stock"}),
            "clientes": 4 * len(h & {"documento", "cedula", "numero_documento", "telefono"}),
            "motos": 5 * len(h & {"placa", "marca", "modelo"}),
            "ventas_factura": 4 * len(h & {"nofactura", "prefijo", "total", "cliente"}),
            "ventas_remision": 4 * len(h & {"noremision", "remision", "total", "cliente"}),
            "ventas_cotizacion": 4 * len(h & {"cotizacion", "total", "cliente"}),
            "detalles_factura": 5 * len(h & {"nofactura", "codigo", "cantidad", "precioventa"}),
            "detalles_remision": 5 * len(h & {"noremision", "remision", "codigo", "cantidad"}),
            "compras": 4 * len(h & {"compra", "cantidad", "codigo", "factura"}),
            "descargas": 4 * len(h & {"descarga", "cantidad", "codigo", "motivo"}),
            "anulaciones_factura": 5 * len(h & {"nofactura", "causa", "motivo"}),
            "anulaciones_remision": 5 * len(h & {"noremision", "remision", "causa"}),
            "usuarios": 4 * len(h & {"usuario", "username", "vendedor", "empleado"}),
            "datosempresa": 4 * len(h & {"razon_social", "nit", "direccion", "telefono"}),
            "auditoria": 5 * len(h & {"accion", "usuario", "fecha", "modelo"}),
            "report_only": 1,
        }

        filename = dataset.path.stem.lower()
        if "view_" in filename or "resumen" in filename or "ivas" in filename:
            scores["report_only"] += 8
        if "prefactura" in filename:
            scores["ventas_cotizacion"] += 4
        if "rem" in filename and "detalle" in filename:
            scores["detalles_remision"] += 3
        if "factura" in filename and "detalle" in filename:
            scores["detalles_factura"] += 3

        if any(any(k in str(v).lower() for k in ["fac", "rem", "cot"]) for row in sample for v in row.values() if v):
            scores["ventas_factura"] += 1
            scores["ventas_remision"] += 1

        return max(scores.items(), key=lambda kv: kv[1])[0]

    def _dispatch(self, dataset: Dataset, fr: FileReport) -> None:
        handlers = {
            "categorias": self._import_categorias,
            "impuestos": self._import_impuestos,
            "productos": self._import_productos,
            "clientes": self._import_clientes,
            "motos": self._import_motos,
            "ventas_factura": lambda d, r: self._import_ventas(d, r, "FACTURA"),
            "ventas_remision": lambda d, r: self._import_ventas(d, r, "REMISION"),
            "ventas_cotizacion": lambda d, r: self._import_ventas(d, r, "COTIZACION"),
            "detalles_factura": lambda d, r: self._import_detalles(d, r, "FACTURA"),
            "detalles_remision": lambda d, r: self._import_detalles(d, r, "REMISION"),
            "compras": self._import_compras,
            "descargas": self._import_descargas,
            "anulaciones_factura": lambda d, r: self._import_anulaciones(d, r, "FACTURA"),
            "anulaciones_remision": lambda d, r: self._import_anulaciones(d, r, "REMISION"),
            "usuarios": self._import_usuarios,
            "datosempresa": self._import_datosempresa,
            "auditoria": self._import_auditoria,
            "report_only": self._archive_only,
        }
        handlers.get(fr.classification, self._archive_only)(dataset, fr)

    def _record_unmapped(self, dataset: Dataset, row: dict[str, Any], reason: str, target: str) -> None:
        self.unmapped_payloads.append(
            {
                "file": dataset.path.name,
                "sheet": dataset.sheet,
                "row": row.get("_row_number"),
                "target": target,
                "reason": reason,
                "payload": {k: v for k, v in row.items() if not k.startswith("_")},
            }
        )

    def _admin_user(self):
        if self.cache["users"].get("admin"):
            return self.cache["users"]["admin"]
        User = get_user_model()
        user = User.objects.filter(is_superuser=True).order_by("id").first() or User.objects.order_by("id").first()
        self.cache["users"]["admin"] = user
        return user

    def _find_user(self, row: dict[str, Any], role_hint: str | None = None):
        User = get_user_model()
        username = self._pick(row, ["username", "usuario", "user", "login"])
        full_name = self._pick(row, ["nombre", "vendedor", "empleado"]) or ""
        key = (username or full_name).lower().strip()
        if key and self.cache["users"].get(key):
            return self.cache["users"][key]

        qs = User.objects
        user = None
        if username:
            user = qs.filter(username__iexact=username).first()
        if not user and full_name:
            bits = full_name.split(maxsplit=1)
            user = qs.filter(first_name__iexact=bits[0]).first()

        if not user:
            seed = username or slug(full_name) or f"legacy_{timezone.now().strftime('%H%M%S')}"
            user, _ = qs.update_or_create(
                username=seed[:150],
                defaults={
                    "first_name": bits[0] if full_name and (bits := full_name.split(maxsplit=1)) else full_name[:150],
                    "last_name": bits[1] if full_name and len(bits) > 1 else "",
                    "email": (self._pick(row, ["email", "correo"]) or "")[:254],
                    "is_active": False,
                    "tipo_usuario": "VENDEDOR" if role_hint == "vendedor" else "MECANICO" if role_hint == "mecanico" else "BODEGUERO",
                },
            )
            user.set_unusable_password()
            user.save(update_fields=["password"])

        if role_hint == "vendedor":
            PerfilVendedor.objects.get_or_create(usuario=user)
        self.cache["users"][key] = user
        return user

    def _pick(self, row: dict[str, Any], keys: list[str], default: Any = None) -> Any:
        for key in keys:
            if key in row and row[key] not in (None, ""):
                return row[key]
        return default

    def _numero_comprobante(self, row: dict[str, Any], tipo: str, idx: int) -> str:
        if tipo == "FACTURA":
            pref = str(self._pick(row, ["prefijo"], "FAC") or "FAC").strip()
            num = self._pick(row, ["nofactura", "factura", "numero", "numero_comprobante"])
            if num:
                return f"{pref}-{str(num).strip()}"
        if tipo == "REMISION":
            num = self._pick(row, ["noremision", "remision", "numero", "numero_comprobante"])
            if num:
                return str(num).strip()
        num = self._pick(row, ["cotizacion", "numero", "numero_comprobante"])
        return str(num).strip() if num else f"{tipo}-{idx}"

    def _doc_key(self, row: dict[str, Any], idx: int, dataset_name: str) -> str:
        for k in DOC_KEYS:
            if row.get(k):
                return str(row[k]).strip()
        return f"LEGACY-{dataset_name}-{idx}"

    def _get_or_create_categoria(self, nombre: str, descripcion: str = "") -> tuple[Categoria, bool]:
        nombre = (nombre or "Sin categoría").strip()
        normalized = nombre.lower()
        if self.cache["categorias"].get(normalized):
            return self.cache["categorias"][normalized], False
        obj, created = Categoria.objects.update_or_create(
            nombre__iexact=nombre,
            defaults={"nombre": nombre, "descripcion": descripcion[:1000], "orden": 0},
        )
        self.cache["categorias"][normalized] = obj
        return obj, created

    def _get_or_create_proveedor(self, row: dict[str, Any]) -> tuple[Proveedor | None, bool]:
        nit = str(self._pick(row, ["nit", "identificacion"], "") or "").strip()
        nombre = str(self._pick(row, ["proveedor", "marca", "fabricante", "nombre_proveedor"], "") or "").strip()
        key = (nit or nombre).lower()
        if not key:
            return None, False
        if self.cache["proveedores"].get(key):
            return self.cache["proveedores"][key], False

        if nit:
            obj, created = Proveedor.objects.update_or_create(
                nit=nit,
                defaults={
                    "nombre": nombre or f"Proveedor {nit}",
                    "telefono": str(self._pick(row, ["telefono", "tel"], "") or ""),
                    "email": str(self._pick(row, ["email", "correo"], "") or ""),
                    "direccion": str(self._pick(row, ["direccion"], "") or ""),
                    "ciudad": str(self._pick(row, ["ciudad", "municipio"], "") or ""),
                },
            )
        else:
            obj, created = Proveedor.objects.update_or_create(
                nombre__iexact=nombre,
                defaults={"nombre": nombre},
            )
        self.cache["proveedores"][key] = obj
        return obj, created

    def _import_categorias(self, dataset: Dataset, fr: FileReport) -> None:
        mapped = {"categoria", "nombre", "descripcion", "detalle"}
        fr.unmapped_columns = [h for h in dataset.headers if h not in mapped and not h.startswith("_")]
        for idx, row in enumerate(dataset.rows, start=1):
            nombre = str(self._pick(row, ["categoria", "nombre", "descripcion"], "") or "").strip()
            if not nombre:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "sin nombre de categoria", "categorias")
                continue
            _, created = self._get_or_create_categoria(nombre, str(self._pick(row, ["descripcion", "detalle"], "") or ""))
            fr.imported += int(created)
            fr.updated += int(not created)

    def _import_impuestos(self, dataset: Dataset, fr: FileReport) -> None:
        mapped = {"impuesto", "nombre", "iva", "porcentaje", "descripcion"}
        fr.unmapped_columns = [h for h in dataset.headers if h not in mapped and not h.startswith("_")]
        for row in dataset.rows:
            nombre = str(self._pick(row, ["impuesto", "nombre", "descripcion"], "") or "").strip()
            if not nombre:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "sin nombre de impuesto", "impuestos")
                continue
            porcentaje = to_decimal(self._pick(row, ["porcentaje", "iva"], 0), Decimal("0"))
            obj, created = Impuesto.objects.update_or_create(
                nombre__iexact=nombre,
                defaults={"nombre": nombre, "porcentaje": porcentaje},
            )
            self.cache["impuestos"][obj.nombre.lower()] = obj
            fr.imported += int(created)
            fr.updated += int(not created)

    def _import_clientes(self, dataset: Dataset, fr: FileReport) -> None:
        mapped = set(DOC_KEYS + NAME_KEYS + ["tipo_documento", "telefono", "celular", "email", "correo", "direccion", "ciudad", "municipio"])
        fr.unmapped_columns = [h for h in dataset.headers if h not in mapped and not h.startswith("_")]
        for idx, row in enumerate(dataset.rows, start=1):
            documento = self._doc_key(row, idx, dataset.path.stem)
            nombre = str(self._pick(row, NAME_KEYS, "Cliente Legacy") or "Cliente Legacy").strip()
            defaults = {
                "tipo_documento": str(self._pick(row, ["tipo_documento", "tipo"], "CC") or "CC")[:50],
                "nombre": nombre[:200],
                "telefono": str(self._pick(row, ["telefono", "celular"], "") or "")[:50],
                "email": str(self._pick(row, ["email", "correo"], "") or "")[:254],
                "direccion": str(self._pick(row, ["direccion", "domicilio"], "") or ""),
                "ciudad": str(self._pick(row, ["ciudad", "municipio"], "") or "")[:100],
            }
            obj, created = Cliente.objects.update_or_create(numero_documento=documento, defaults=defaults)
            self.cache["clientes"][documento] = obj
            self.cache["clientes_nombre"][obj.nombre.lower()] = obj
            fr.imported += int(created)
            fr.updated += int(not created)

    def _import_usuarios(self, dataset: Dataset, fr: FileReport) -> None:
        for row in dataset.rows:
            role = "vendedor" if any(k in dataset.path.stem.lower() for k in ["vendedor", "usuario"]) else None
            user = self._find_user(row, role_hint=role)
            if user:
                fr.updated += 1
            else:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "no se pudo crear usuario", "usuarios")

    def _import_productos(self, dataset: Dataset, fr: FileReport) -> None:
        mapped = {
            "codigo", "cod", "referencia", "id", "articulo", "nombre", "descripcion", "detalle", "categoria", "linea",
            "proveedor", "marca", "fabricante", "costo", "precio_costo", "compra", "precio_venta", "precioventa", "valor",
            "stock", "existencias", "cantidad", "unidad", "unidad_medida", "iva", "porcentaje",
        }
        fr.unmapped_columns = [h for h in dataset.headers if h not in mapped and not h.startswith("_")]
        default_cat, _ = self._get_or_create_categoria("Sin categoría", "Asignada durante importación legacy")
        for idx, row in enumerate(dataset.rows, start=1):
            codigo = str(self._pick(row, ["codigo", "cod", "referencia", "id"], "") or "").strip()
            nombre = str(self._pick(row, ["nombre", "articulo", "descripcion"], "") or "").strip()
            if not codigo or not nombre:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "producto sin codigo o nombre", "productos")
                continue

            cat_name = str(self._pick(row, ["categoria", "linea", "grupo"], "") or "").strip()
            categoria, _ = self._get_or_create_categoria(cat_name, "") if cat_name else (default_cat, False)
            proveedor, _ = self._get_or_create_proveedor(row)

            costo = to_decimal(self._pick(row, ["costo", "precio_costo", "compra"], 1), Decimal("1"))
            precio = to_decimal(self._pick(row, ["precio_venta", "precioventa", "valor", "precio"], 1), Decimal("1"))
            stock = to_decimal(self._pick(row, ["stock", "existencias", "cantidad"], 0), Decimal("0"))
            iva = to_decimal(self._pick(row, ["iva", "porcentaje"], 19), Decimal("19"))
            unidad = str(self._pick(row, ["unidad", "unidad_medida"], "N/A") or "N/A").upper()
            if unidad not in {"N/A", "KG", "LT", "MT"}:
                unidad = "N/A"

            defaults = {
                "nombre": nombre[:300],
                "descripcion": str(self._pick(row, ["descripcion", "detalle"], "") or ""),
                "categoria": categoria,
                "proveedor": proveedor,
                "precio_costo": max(costo, Decimal("0.01")),
                "precio_venta": max(precio, Decimal("0.01")),
                "precio_venta_minimo": max(precio, Decimal("0.01")),
                "stock": stock,
                "stock_minimo": Decimal("5"),
                "unidad_medida": unidad,
                "iva_porcentaje": iva,
                "iva_exento": iva == 0,
                "aplica_descuento": True,
                "es_servicio": False,
            }
            obj, created = Producto.objects.update_or_create(codigo=codigo, defaults=defaults)
            self.cache["productos"][obj.codigo] = obj
            self.cache["productos_nombre"][obj.nombre.lower()] = obj
            fr.imported += int(created)
            fr.updated += int(not created)

    def _import_motos(self, dataset: Dataset, fr: FileReport) -> None:
        for row in dataset.rows:
            placa = str(self._pick(row, ["placa", "moto", "matricula"], "") or "").strip().upper()
            if not placa:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "moto sin placa", "motos")
                continue
            cliente = None
            doc = self._pick(row, DOC_KEYS)
            if doc:
                cliente = self.cache["clientes"].get(str(doc).strip()) or Cliente.objects.filter(numero_documento=str(doc).strip()).first()
            mecanico_name = self._pick(row, ["mecanico", "empleado"])
            mecanico = None
            if mecanico_name:
                mecanico, _ = Mecanico.objects.get_or_create(nombre=str(mecanico_name).strip())
            proveedor, _ = self._get_or_create_proveedor(row)

            defaults = {
                "marca": str(self._pick(row, ["marca"], "") or "")[:100],
                "modelo": str(self._pick(row, ["modelo"], "") or "")[:100],
                "color": str(self._pick(row, ["color"], "") or "")[:50],
                "anio": int(to_decimal(self._pick(row, ["anio", "year"], 0), Decimal("0"))) or None,
                "cliente": cliente,
                "mecanico": mecanico,
                "proveedor": proveedor,
                "fecha_ingreso": (to_dt(self._pick(row, ["fecha_ingreso", "fecha"], None)) or timezone.now()).date(),
                "observaciones": str(self._pick(row, ["observaciones", "nota"], "") or ""),
            }
            _, created = Moto.objects.update_or_create(placa=placa, defaults=defaults)
            fr.imported += int(created)
            fr.updated += int(not created)

    def _get_cliente(self, row: dict[str, Any], idx: int, dataset_name: str) -> Cliente:
        doc = self._doc_key(row, idx, dataset_name)
        obj = self.cache["clientes"].get(doc) or Cliente.objects.filter(numero_documento=doc).first()
        if obj:
            return obj
        nombre = str(self._pick(row, NAME_KEYS, "Cliente Legacy") or "Cliente Legacy")
        obj, _ = Cliente.objects.get_or_create(numero_documento=doc, defaults={"nombre": nombre, "tipo_documento": "CC"})
        self.cache["clientes"][doc] = obj
        return obj

    def _import_ventas(self, dataset: Dataset, fr: FileReport, tipo: str) -> None:
        for idx, row in enumerate(dataset.rows, start=1):
            numero = self._numero_comprobante(row, tipo, idx)
            cliente = self._get_cliente(row, idx, dataset.path.stem)
            vendedor = self._find_user(row, role_hint="vendedor") or self._admin_user()
            subtotal = to_decimal(self._pick(row, ["subtotal"], 0), Decimal("0"))
            descuento = to_decimal(self._pick(row, ["descuento", "descuento_valor"], 0), Decimal("0"))
            iva = to_decimal(self._pick(row, ["iva", "impuesto"], 0), Decimal("0"))
            total = to_decimal(self._pick(row, ["total", "valor"], subtotal + iva - descuento), subtotal + iva - descuento)
            medio = str(self._pick(row, ["medio_pago", "mediopago", "pago", "forma_pago"], "EFECTIVO") or "EFECTIVO").upper()
            estado = str(self._pick(row, ["estado"], "FACTURADA") or "FACTURADA").upper()
            if medio not in {m[0] for m in Venta.MEDIO_PAGO}:
                medio = "EFECTIVO"
            if estado not in {e[0] for e in Venta.ESTADO}:
                estado = "FACTURADA"
            fecha = to_dt(self._pick(row, ["fecha", "fecha_hora"], None))

            defaults = {
                "tipo_comprobante": tipo,
                "cliente": cliente,
                "vendedor": vendedor,
                "subtotal": subtotal,
                "descuento_porcentaje": Decimal("0"),
                "descuento_valor": descuento,
                "iva": iva,
                "total": total,
                "medio_pago": medio,
                "efectivo_recibido": total,
                "cambio": Decimal("0"),
                "estado": estado,
                "observaciones": str(self._pick(row, ["observaciones", "nota"], "") or ""),
            }
            venta, created = Venta.objects.update_or_create(numero_comprobante=numero, defaults=defaults)
            if fecha:
                Venta.objects.filter(pk=venta.pk).update(fecha=fecha)
            if "prefactura" in dataset.path.stem.lower():
                fr.notes.append("PreFactura se migró como COTIZACION para trazabilidad operativa.")
            self.cache["ventas"][f"{tipo}|{numero}"] = venta
            fr.imported += int(created)
            fr.updated += int(not created)

    def _find_producto(self, row: dict[str, Any]) -> Producto | None:
        codigo = str(self._pick(row, ["codigo", "cod", "referencia", "id"], "") or "").strip()
        if codigo:
            obj = self.cache["productos"].get(codigo) or Producto.objects.filter(codigo=codigo).first()
            if obj:
                self.cache["productos"][codigo] = obj
                return obj
        nombre = str(self._pick(row, ["producto", "articulo", "nombre", "descripcion"], "") or "").strip()
        if nombre:
            return self.cache["productos_nombre"].get(nombre.lower()) or Producto.objects.filter(nombre__iexact=nombre).first()
        return None

    def _import_detalles(self, dataset: Dataset, fr: FileReport, tipo: str) -> None:
        for idx, row in enumerate(dataset.rows, start=1):
            numero = self._numero_comprobante(row, tipo, idx)
            venta = self.cache["ventas"].get(f"{tipo}|{numero}") or Venta.objects.filter(tipo_comprobante=tipo, numero_comprobante=numero).first()
            producto = self._find_producto(row)
            if not venta or not producto:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "detalle sin venta o producto", "detalles_venta")
                continue
            cantidad = to_decimal(self._pick(row, ["cantidad", "cant"], 1), Decimal("1"))
            precio = to_decimal(self._pick(row, ["precioventa", "pventau", "preciou", "precio", "precio_unitario"], 1), Decimal("1"))
            descuento = to_decimal(self._pick(row, ["descuento", "descu_valor", "descuento_unitario"], 0), Decimal("0"))
            iva = to_decimal(self._pick(row, ["iva", "porcentaje"], producto.iva_porcentaje), producto.iva_porcentaje)
            subtotal = (precio * cantidad).quantize(Decimal("0.01"))
            total = (subtotal - (descuento * cantidad)).quantize(Decimal("0.01"))

            detail_key = {
                "venta": venta,
                "producto": producto,
                "cantidad": cantidad,
                "precio_unitario": precio,
                "descuento_unitario": descuento,
                "iva_porcentaje": iva,
            }
            detalle, created = DetalleVenta.objects.update_or_create(defaults={"subtotal": subtotal, "total": total, "afecto_inventario": True}, **detail_key)
            if detalle:
                fr.imported += int(created)
                fr.updated += int(not created)

    def _import_anulaciones(self, dataset: Dataset, fr: FileReport, tipo: str) -> None:
        admin = self._admin_user()
        for idx, row in enumerate(dataset.rows, start=1):
            numero = self._numero_comprobante(row, tipo, idx)
            venta = self.cache["ventas"].get(f"{tipo}|{numero}") or Venta.objects.filter(numero_comprobante=numero, tipo_comprobante=tipo).first()
            if not venta:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "anulacion sin documento origen", "ventas_anuladas/remisiones_anuladas")
                continue
            motivo = str(self._pick(row, ["motivo", "causa", "razon"], "OTRO") or "OTRO").upper()
            if motivo not in dict(VentaAnulada.MOTIVO_CHOICES):
                motivo = "OTRO"
            descripcion = str(self._pick(row, ["descripcion", "detalle", "observacion", "causa"], "Anulación legacy") or "Anulación legacy")

            model = VentaAnulada if tipo == "FACTURA" else RemisionAnulada
            key = {"venta": venta} if tipo == "FACTURA" else {"remision": venta}
            _, created = model.objects.update_or_create(
                **key,
                defaults={"motivo": motivo, "descripcion": descripcion, "anulado_por": admin, "devuelve_inventario": True},
            )
            fr.imported += int(created)
            fr.updated += int(not created)

    def _import_compras(self, dataset: Dataset, fr: FileReport) -> None:
        admin = self._admin_user()
        for row in dataset.rows:
            producto = self._find_producto(row)
            if not producto:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "compra sin producto", "movimientos_inventario")
                continue
            cantidad = to_decimal(self._pick(row, ["cantidad", "cant"], 0), Decimal("0"))
            if cantidad <= 0:
                fr.rejected += 1
                continue
            costo = to_decimal(self._pick(row, ["compra", "precio_compra", "costo", "valor"], producto.precio_costo), producto.precio_costo)
            referencia = str(self._pick(row, ["factura", "documento", "referencia"], "COMPRA") or "COMPRA")
            stock_anterior = producto.stock
            stock_nuevo = stock_anterior + cantidad
            _, created = MovimientoInventario.objects.update_or_create(
                producto=producto,
                tipo="ENTRADA",
                cantidad=cantidad,
                referencia=referencia,
                defaults={
                    "stock_anterior": stock_anterior,
                    "stock_nuevo": stock_nuevo,
                    "costo_unitario": costo,
                    "usuario": admin,
                    "observaciones": str(self._pick(row, ["observaciones", "nota", "proveedor"], "") or ""),
                },
            )
            producto.stock = stock_nuevo
            producto.precio_costo = costo
            producto.ultima_compra = timezone.now()
            producto.save(update_fields=["stock", "precio_costo", "ultima_compra"], touch_ultima_compra=False)
            fr.imported += int(created)
            fr.updated += int(not created)

    def _import_descargas(self, dataset: Dataset, fr: FileReport) -> None:
        admin = self._admin_user()
        for row in dataset.rows:
            producto = self._find_producto(row)
            if not producto:
                fr.rejected += 1
                self._record_unmapped(dataset, row, "descarga sin producto", "movimientos_inventario")
                continue
            cantidad = to_decimal(self._pick(row, ["cantidad", "cant"], 0), Decimal("0"))
            if cantidad <= 0:
                fr.rejected += 1
                continue
            motivo = str(self._pick(row, ["motivo", "tipo", "concepto"], "SALIDA") or "SALIDA").upper()
            mov_type = "BAJA" if "BAJA" in motivo else "AJUSTE" if "AJUSTE" in motivo else "SALIDA"
            stock_anterior = producto.stock
            stock_nuevo = stock_anterior - cantidad
            _, created = MovimientoInventario.objects.update_or_create(
                producto=producto,
                tipo=mov_type,
                cantidad=-cantidad,
                referencia=str(self._pick(row, ["documento", "referencia"], "DESCARGA") or "DESCARGA"),
                defaults={
                    "stock_anterior": stock_anterior,
                    "stock_nuevo": stock_nuevo,
                    "costo_unitario": producto.precio_costo,
                    "usuario": admin,
                    "observaciones": str(self._pick(row, ["observaciones", "nota"], "") or ""),
                },
            )
            producto.stock = stock_nuevo
            producto.save(update_fields=["stock"], touch_ultima_compra=False)
            fr.imported += int(created)
            fr.updated += int(not created)

    def _import_datosempresa(self, dataset: Dataset, fr: FileReport) -> None:
        if not dataset.rows:
            return
        row = dataset.rows[0]
        empresa_defaults = {
            "tipo_identificacion": "NIT",
            "identificacion": str(self._pick(row, ["nit", "identificacion"], "") or ""),
            "dv": str(self._pick(row, ["dv"], "") or "")[:1],
            "tipo_persona": "Persona jurídica",
            "razon_social": str(self._pick(row, ["razon_social", "empresa", "nombre"], "Empresa Legacy") or "Empresa Legacy"),
            "regimen": "RÉGIMEN COMÚN",
            "direccion": str(self._pick(row, ["direccion"], "") or ""),
            "ciudad": str(self._pick(row, ["ciudad"], "") or ""),
            "municipio": str(self._pick(row, ["municipio", "ciudad"], "") or ""),
            "telefono": str(self._pick(row, ["telefono"], "") or ""),
            "sitio_web": str(self._pick(row, ["web", "sitio_web"], "") or ""),
            "correo": str(self._pick(row, ["correo", "email"], "") or ""),
        }
        obj = ConfiguracionEmpresa.objects.order_by("id").first()
        if obj:
            for k, v in empresa_defaults.items():
                setattr(obj, k, v)
            obj.save()
            fr.updated += 1
        else:
            ConfiguracionEmpresa.objects.create(**empresa_defaults)
            fr.imported += 1

        cf = ConfiguracionFacturacion.objects.order_by("id").first()
        pref = str(self._pick(row, ["prefijo", "prefijo_factura"], "FAC") or "FAC")
        num = int(to_decimal(self._pick(row, ["numero", "numero_factura"], 1), Decimal("1")))
        if cf:
            cf.prefijo_factura = pref
            cf.numero_factura = num
            cf.save(update_fields=["prefijo_factura", "numero_factura"])
            fr.updated += 1
        else:
            ConfiguracionFacturacion.objects.create(prefijo_factura=pref, numero_factura=num)
            fr.imported += 1

    def _import_auditoria(self, dataset: Dataset, fr: FileReport) -> None:
        for row in dataset.rows:
            accion = str(self._pick(row, ["accion", "tipo"], "OTRO") or "OTRO").upper()
            if accion not in dict(Auditoria.ACCION_CHOICES):
                accion = "OTRO"
            usuario = self._find_user(row)
            fecha = to_dt(self._pick(row, ["fecha", "fecha_hora"], None)) or timezone.now()
            _, created = Auditoria.objects.update_or_create(
                fecha_hora=fecha,
                usuario_nombre=str(self._pick(row, ["usuario", "usuario_nombre", "nombre"], "legacy") or "legacy")[:150],
                accion=accion,
                modelo=str(self._pick(row, ["modelo", "tabla"], "") or "")[:100],
                objeto_id=str(self._pick(row, ["objeto_id", "registro"], "") or "")[:100],
                defaults={
                    "usuario": usuario,
                    "notas": str(self._pick(row, ["notas", "descripcion", "detalle"], "Evento legado") or "Evento legado"),
                    "ip_address": self._pick(row, ["ip", "ip_address"], None),
                },
            )
            fr.imported += int(created)
            fr.updated += int(not created)

    def _archive_only(self, dataset: Dataset, fr: FileReport) -> None:
        fr.ambiguous = len(dataset.rows)
        fr.notes.append("Archivo tratado como reporte/vista derivada; no se importa a tablas operativas.")
        for row in dataset.rows:
            self._record_unmapped(dataset, row, "archivo derivado o sin destino relacional 1:1", "archive/report")

    def _run_validations(self) -> dict[str, Any]:
        details_without_sale = DetalleVenta.objects.filter(venta__isnull=True).count()
        details_without_product = DetalleVenta.objects.filter(producto__isnull=True).count()
        ventas_without_cliente = Venta.objects.filter(cliente__isnull=True).count()
        products_without_category = Producto.objects.filter(categoria__isnull=True).count()
        motos_with_invalid_refs = Moto.objects.filter(cliente__isnull=True, mecanico__isnull=True, proveedor__isnull=True).count()
        movimientos_without_product = MovimientoInventario.objects.filter(producto__isnull=True).count()
        anuladas_without_origin = VentaAnulada.objects.filter(venta__isnull=True).count() + RemisionAnulada.objects.filter(remision__isnull=True).count()
        cabecera_detalle_inconsistente = (
            DetalleVenta.objects.values("venta_id").annotate(total=Count("id")).filter(total=0).count()
        )
        total_legacy_reported = sum(r.rows_read for r in self.reports if "ventas_" in r.classification)
        total_imported_ventas = Venta.objects.count()

        return {
            "entity_counts": {
                "categorias": Categoria.objects.count(),
                "impuestos": Impuesto.objects.count(),
                "proveedores": Proveedor.objects.count(),
                "usuarios": get_user_model().objects.count(),
                "mecanicos": Mecanico.objects.count(),
                "clientes": Cliente.objects.count(),
                "productos": Producto.objects.count(),
                "motos": Moto.objects.count(),
                "ventas": total_imported_ventas,
                "detalles_venta": DetalleVenta.objects.count(),
                "movimientos_inventario": MovimientoInventario.objects.count(),
                "ventas_anuladas": VentaAnulada.objects.count(),
                "remisiones_anuladas": RemisionAnulada.objects.count(),
                "auditoria": Auditoria.objects.count(),
            },
            "checks": {
                "ventas_sin_cliente": ventas_without_cliente,
                "detalles_sin_venta": details_without_sale,
                "detalles_sin_producto": details_without_product,
                "productos_sin_categoria": products_without_category,
                "motos_sin_referencias": motos_with_invalid_refs,
                "movimientos_sin_producto": movimientos_without_product,
                "documentos_anulados_sin_origen": anuladas_without_origin,
                "inconsistencias_cabecera_detalle": cabecera_detalle_inconsistente,
                "comparacion_totales_legacy_vs_importado": {
                    "legacy_rows_ventas": total_legacy_reported,
                    "ventas_importadas_db": total_imported_ventas,
                    "diferencia": total_imported_ventas - total_legacy_reported,
                },
            },
            "files_summary": {
                "total_files": len(self.reports),
                "fully_imported": [r.filename for r in self.reports if r.rows_read and r.rejected == 0 and r.ambiguous == 0],
                "partial": [r.filename for r in self.reports if r.rejected > 0],
                "ambiguous": [r.filename for r in self.reports if r.ambiguous > 0],
            },
        }

    def _cleanup_staging_tables(self, validations: dict[str, Any]) -> dict[str, Any]:
        critical_errors = any(validations["checks"][k] > 0 for k in [
            "detalles_sin_venta", "detalles_sin_producto", "movimientos_sin_producto", "documentos_anulados_sin_origen"
        ])
        result = {"enabled": self.cleanup_temp_on_success, "dropped": [], "skipped_reason": ""}
        if not self.cleanup_temp_on_success:
            result["skipped_reason"] = "flag --cleanup-temp-on-success no activado"
            return result
        if not self.commit:
            result["skipped_reason"] = "modo dry-run"
            return result
        if critical_errors:
            result["skipped_reason"] = "validaciones críticas fallidas"
            return result

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema='public' AND table_name LIKE 'staging_%'
                """
            )
            names = [r[0] for r in cursor.fetchall()]
            for name in names:
                cursor.execute(f'DROP TABLE IF EXISTS "{name}" CASCADE')
                result["dropped"].append(name)
        return result

    def _persist_reports(self, payload: dict[str, Any]) -> dict[str, Any]:
        out_dir = self.base_path / "migration_reports"
        out_dir.mkdir(parents=True, exist_ok=True)
        ts = timezone.now().strftime("%Y%m%d_%H%M%S")
        json_path = out_dir / f"legacy_import_report_{ts}.json"
        md_path = out_dir / f"legacy_import_report_{ts}.md"

        json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

        lines = [
            "# Reporte de importación legacy XLSX",
            "",
            f"- Fecha: {payload['timestamp']}",
            f"- Modo: {payload['mode']}",
            f"- Ruta: `{payload['path']}`",
            "",
            "## Archivos procesados",
            "",
            "| Archivo | Clasificación | Leídas | Importadas | Actualizadas | Rechazadas | Ambiguas |",
            "|---|---:|---:|---:|---:|---:|---:|",
        ]
        for item in payload["files"]:
            lines.append(
                f"| {item['filename']} | {item['classification']} | {item['rows_read']} | {item['imported']} | {item['updated']} | {item['rejected']} | {item['ambiguous']} |"
            )
        lines.extend([
            "",
            "## Validaciones",
            "",
            "```json",
            json.dumps(payload["validations"], ensure_ascii=False, indent=2),
            "```",
            "",
            "## Limpieza staging",
            "",
            "```json",
            json.dumps(payload["cleanup"], ensure_ascii=False, indent=2),
            "```",
            "",
            f"## Datos preservados fuera de tablas reales ({len(payload['unmapped_payloads'])})",
        ])
        md_path.write_text("\n".join(lines), encoding="utf-8")
        payload["report_files"] = [str(json_path), str(md_path)]
        return payload
