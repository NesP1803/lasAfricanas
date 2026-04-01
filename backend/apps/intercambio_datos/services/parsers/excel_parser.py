from openpyxl import load_workbook
import xlrd


def parse_excel(file_path, extension):
    if extension in {'.xls'}:
        wb = xlrd.open_workbook(file_path)
        sheets = []
        for s in wb.sheets():
            if s.nrows == 0:
                continue
            headers = [str(s.cell_value(0, c)).strip() for c in range(s.ncols)]
            rows = []
            for r in range(1, s.nrows):
                values = [s.cell_value(r, c) for c in range(s.ncols)]
                if not any(str(v).strip() for v in values):
                    continue
                rows.append(dict(zip(headers, values)))
            sheets.append({'sheet_name': s.name, 'headers': headers, 'rows': rows})
        return sheets

    wb = load_workbook(file_path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        headers = [str(v).strip() if v is not None else '' for v in all_rows[0]]
        rows = []
        for row in all_rows[1:]:
            if not any(row):
                continue
            values = [str(v).strip() if v is not None else '' for v in row]
            rows.append(dict(zip(headers, values)))
        sheets.append({'sheet_name': ws.title, 'headers': headers, 'rows': rows})
    return sheets
